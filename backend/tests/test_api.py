from __future__ import annotations

import io
import os
import shutil
import tempfile
from types import SimpleNamespace
from atexit import register
from pathlib import Path

TEST_DATA_DIR = Path(tempfile.mkdtemp(prefix="industry-atlas-tests-"))
os.environ["DATABASE_PATH"] = str(TEST_DATA_DIR / "test.db")
os.environ["UPLOAD_DIR"] = str(TEST_DATA_DIR / "uploads")
# Tests must never inherit production credentials or contact paid external APIs.
os.environ["DEEPSEEK_API_KEY"] = ""
os.environ["WECHAT_APP_ID"] = ""
os.environ["WECHAT_APP_SECRET"] = ""
register(lambda: shutil.rmtree(TEST_DATA_DIR, ignore_errors=True))

from fastapi.testclient import TestClient
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfWriter
from reportlab.pdfgen import canvas

from app.main import app
from app import main as main_module
from app.extraction import validate_findings
from app import documents as document_module


def make_pdf() -> bytes:
    output = io.BytesIO()
    writer = PdfWriter()
    writer.add_blank_page(width=300, height=300)
    writer.write(output)
    return output.getvalue()


def make_table_pdf() -> bytes:
    output = io.BytesIO()
    pdf = canvas.Canvas(output, pagesize=(400, 400))
    x_positions, y_positions = [40, 150, 260, 360], [320, 280, 240, 200]
    for x in x_positions:
        pdf.line(x, y_positions[-1], x, y_positions[0])
    for y in y_positions:
        pdf.line(x_positions[0], y, x_positions[-1], y)
    values = [["Metric", "2025", "2024"], ["Revenue", "100", "80"], ["Growth", "25%", "10%"]]
    for row, values_row in enumerate(values):
        for column, value in enumerate(values_row):
            pdf.drawString(x_positions[column] + 6, y_positions[row + 1] + 14, value)
    pdf.save()
    return output.getvalue()


def test_project_and_document_flow() -> None:
    with TestClient(app) as client:
        assert client.get("/api/health").json()["status"] == "ok"
        login = client.post("/api/auth/dev-login", json={"display_name": "研究员", "device_key": "test-device-key-0001"})
        assert login.status_code == 200
        session = login.json()
        headers = {"Authorization": f"Bearer {session['token']}"}
        workspace_id = session["workspace_id"]
        repeated_login = client.post("/api/auth/dev-login", json={"display_name": "研究员（已返回）", "device_key": "test-device-key-0001"})
        assert repeated_login.status_code == 200
        assert repeated_login.json()["user"]["id"] == session["user"]["id"]
        assert repeated_login.json()["workspace_id"] == workspace_id

        created = client.post(
            "/api/projects",
            headers=headers,
            json={"workspace_id": workspace_id, "title": "光伏产业链", "state": {"nodes": {"industry": {"title": "光伏产业", "summary": "光伏发电产业链", "why": "提供清洁能源", "category": "产业总览", "status": "evidenced", "children": []}}, "rootId": "industry", "expanded": ["industry"], "edges": [], "evidenceData": {"industry": [{"type": "ANNUAL", "title": "示例年报", "location": "第 10 页", "quote": "公司从事光伏组件业务", "verified": False}]}, "researchTasks": []}},
        )
        assert created.status_code == 201
        assert "state_json" not in created.json()
        project_id = created.json()["id"]
        export_state = created.json()["state"]
        export_state["companyData"] = [{"id": "company_1", "name": "示例光伏公司", "reportPeriod": "2025年度", "summary": "主营光伏组件", "documents": [], "findings": [], "mappings": [{"id": "map_1", "nodeId": "industry", "status": "confirmed", "score": 96, "reason": "主营业务与产业节点一致"}]}]
        export_state["companyData"][0]["periods"] = [{"period": "2025年度", "summary": "收入增长", "documents": [], "findingIds": ["finding_1"]}]
        export_state["companyData"][0]["findings"] = [{"id": "finding_1", "category": "financial", "category_label": "经营数据", "title": "营业收入", "value": "100亿元", "quote": "实现营业收入100亿元", "reportPeriod": "2025年度", "matched_pages": [20], "citation_status": "matched"}]
        assert client.put(f"/api/projects/{project_id}", headers=headers, json={"state": export_state}).status_code == 200

        company_sync = client.post(
            "/api/companies/sync",
            headers=headers,
            json={"workspace_id": workspace_id, "project_id": project_id, "companies": [{**export_state["companyData"][0], "mappings": [{**export_state["companyData"][0]["mappings"][0], "nodeTitle": "光伏产业", "projectId": project_id}]}]},
        )
        assert company_sync.status_code == 200
        assert company_sync.json()["created"] == 1
        assert company_sync.json()["companies"][0]["data"]["sourceProjects"] == [project_id]
        company_id = company_sync.json()["companies"][0]["id"]

        merged_company = client.post(
            "/api/companies/sync",
            headers=headers,
            json={"workspace_id": workspace_id, "project_id": project_id, "companies": [{"name": " 示例光伏公司 ", "reportPeriod": "2024年度", "documents": [], "findings": [{"category": "financial", "title": "营业收入", "value": "80亿元", "reportPeriod": "2024年度"}], "periods": [{"period": "2024年度", "documents": [], "findingIds": []}], "mappings": []}]},
        )
        assert merged_company.status_code == 200
        assert merged_company.json()["updated"] == 1
        assert len(merged_company.json()["companies"][0]["data"]["periods"]) == 2
        assert client.get(f"/api/companies/{company_id}", headers=headers).json()["name"] == "示例光伏公司"
        assert len(client.get(f"/api/companies?workspace_id={workspace_id}&q=组件", headers=headers).json()) == 1

        docx_export = client.get(f"/api/projects/{project_id}/export?format=docx", headers=headers)
        assert docx_export.status_code == 200
        assert "wordprocessingml" in docx_export.headers["content-type"]
        assert "光伏产业" in "\n".join(paragraph.text for paragraph in Document(io.BytesIO(docx_export.content)).paragraphs)

        xlsx_export = client.get(f"/api/projects/{project_id}/export?format=xlsx", headers=headers)
        assert xlsx_export.status_code == 200
        workbook = load_workbook(io.BytesIO(xlsx_export.content), read_only=True)
        assert workbook.sheetnames == ["节点", "关系", "证据", "任务", "公司", "公司映射", "公司期间", "公司指标", "年报表格"]
        assert workbook["节点"]["B2"].value == "光伏产业"
        assert workbook["公司映射"]["B2"].value == "示例光伏公司"
        assert workbook["公司指标"]["F2"].value == "100亿元"

        pdf_export = client.get(f"/api/projects/{project_id}/export?format=pdf", headers=headers)
        assert pdf_export.status_code == 200
        assert pdf_export.content.startswith(b"%PDF")

        imported = client.post(
            "/api/projects/import",
            headers=headers,
            json={"workspace_id": workspace_id, "state": {"projectTitle": "旧项目", "nodes": {"industry": {}}, "rootId": "industry"}},
        )
        assert imported.status_code == 201

        upload = client.post(
            f"/api/documents?workspace_id={workspace_id}&project_id={project_id}",
            headers={**headers, "X-Filename": "annual-report.pdf", "Content-Type": "application/pdf"},
            content=make_pdf(),
        )
        assert upload.status_code == 202
        document = client.get(f"/api/documents/{upload.json()['id']}", headers=headers)
        assert document.status_code == 200
        assert document.json()["status"] == "ready"
        assert document.json()["page_count"] == 1
        assert document.json()["needs_ocr"] == 1
        assert document.json()["table_count"] == 0

        tables = client.get(f"/api/documents/{upload.json()['id']}/tables", headers=headers)
        assert tables.status_code == 200
        assert tables.json() == []

        class FakeOcr:
            def __call__(self, _: bytes) -> SimpleNamespace:
                return SimpleNamespace(txts=("扫描年报主营业务", "营业收入持续增长"))

        document_module._ocr_engine = FakeOcr()
        ocr = client.post(f"/api/documents/{upload.json()['id']}/ocr", headers=headers)
        assert ocr.status_code == 202
        ocr_document = client.get(f"/api/documents/{upload.json()['id']}", headers=headers).json()
        assert ocr_document["status"] == "ready"
        assert ocr_document["extraction_mode"] == "ocr"
        assert ocr_document["ocr_page_count"] == 1
        assert ocr_document["char_count"] > 0

        documents = client.get(
            f"/api/documents?workspace_id={workspace_id}&project_id={project_id}",
            headers=headers,
        )
        assert documents.status_code == 200
        assert documents.json()[0]["filename"] == "annual-report.pdf"
        assert documents.json()[0]["extraction_count"] == 0

        extraction_history = client.get(
            f"/api/documents/{upload.json()['id']}/extractions",
            headers=headers,
        )
        assert extraction_history.status_code == 200
        assert extraction_history.json() == []

        retried = client.post(f"/api/documents/{upload.json()['id']}/retry", headers=headers)
        assert retried.status_code == 202
        assert client.get(f"/api/documents/{upload.json()['id']}", headers=headers).json()["status"] == "ready"

        table_upload = client.post(
            f"/api/documents?workspace_id={workspace_id}&project_id={project_id}",
            headers={**headers, "X-Filename": "financial-table.pdf", "Content-Type": "application/pdf"},
            content=make_table_pdf(),
        )
        table_document = client.get(f"/api/documents/{table_upload.json()['id']}", headers=headers).json()
        assert table_document["table_count"] == 1
        extracted_tables = client.get(f"/api/documents/{table_upload.json()['id']}/tables", headers=headers).json()
        assert extracted_tables[0]["data"][1] == ["Revenue", "100", "80"]
        table_export = client.get(f"/api/projects/{project_id}/export?format=xlsx", headers=headers)
        table_workbook = load_workbook(io.BytesIO(table_export.content), read_only=True)
        assert table_workbook["年报表格"]["E3"].value == "Revenue"

        ai_without_server_key = client.post(
            "/api/ai/structured",
            headers=headers,
            json={"workspace_id": workspace_id, "project_id": project_id, "system": "只返回 JSON", "user": "返回测试结果"},
        )
        assert ai_without_server_key.status_code == 503
        extraction_without_server_key = client.post(
            f"/api/documents/{upload.json()['id']}/extractions",
            headers=headers,
            json={},
        )
        assert extraction_without_server_key.status_code == 503
        assert client.delete(f"/api/companies/{company_id}", headers=headers).status_code == 204
        assert client.get(f"/api/companies/{company_id}", headers=headers).status_code == 404


def test_permissions_require_login() -> None:
    with TestClient(app) as client:
        assert client.get("/api/workspaces").status_code == 401
        assert client.get("/api/workspaces", headers={"Authorization": "Bearer not-a-token"}).status_code == 401
        assert client.post("/api/auth/wechat", json={"code": "test-code"}).status_code == 503


def test_production_readiness_guard(monkeypatch) -> None:
    blocked = SimpleNamespace(production_issues=lambda: ["missing production config"], app_env="production")
    monkeypatch.setattr(main_module, "settings", blocked)
    with TestClient(app) as client:
        response = client.get("/api/ready")
        assert response.status_code == 503
        assert response.json()["detail"]["status"] == "not_ready"

    ready = SimpleNamespace(production_issues=lambda: [], app_env="production")
    monkeypatch.setattr(main_module, "settings", ready)
    with TestClient(app) as client:
        response = client.get("/api/ready")
        assert response.status_code == 200
        assert response.json() == {"status": "ready", "environment": "production"}


def test_accounts_members_and_project_conflicts() -> None:
    with TestClient(app) as client:
        owner = client.post("/api/auth/register", json={
            "email": "owner@example.com", "password": "correct horse battery",
            "display_name": "负责人", "workspace_name": "新能源研究组",
        })
        assert owner.status_code == 201
        owner_session = owner.json()
        workspace_id = owner_session["workspace"]["id"]
        owner_headers = {"Authorization": f"Bearer {owner_session['token']}"}
        assert client.post("/api/auth/register", json={
            "email": "OWNER@example.com", "password": "another secure password",
            "display_name": "重复账号", "workspace_name": "重复空间",
        }).status_code == 409
        assert client.post("/api/auth/login", json={"email": "owner@example.com", "password": "wrong password"}).status_code == 401
        login = client.post("/api/auth/login", json={"email": "OWNER@example.com", "password": "correct horse battery"})
        assert login.status_code == 200
        assert login.json()["workspace"]["id"] == workspace_id

        editor = client.post("/api/auth/register", json={
            "email": "editor@example.com", "password": "editor password 123",
            "display_name": "协作研究员", "workspace_name": "个人空间",
        }).json()
        editor_headers = {"Authorization": f"Bearer {editor['token']}"}
        editor_id = editor["user"]["id"]
        added = client.post(f"/api/workspaces/{workspace_id}/members", headers=owner_headers, json={"email": "EDITOR@example.com", "role": "editor"})
        assert added.status_code == 201
        assert added.json()["role"] == "editor"
        members = client.get(f"/api/workspaces/{workspace_id}/members", headers=editor_headers)
        assert members.status_code == 200
        assert [member["role"] for member in members.json()] == ["owner", "editor"]
        editor_workspaces = client.get("/api/workspaces", headers=editor_headers)
        assert editor_workspaces.status_code == 200
        assert {workspace["id"] for workspace in editor_workspaces.json()} == {workspace_id, editor["workspace"]["id"]}

        created = client.post("/api/projects", headers=owner_headers, json={
            "workspace_id": workspace_id, "title": "储能产业链", "state": {"projectTitle": "储能产业链", "nodes": {}, "rootId": "industry"},
        })
        assert created.status_code == 201
        project = created.json()
        assert project["version"] == 1
        project_list = client.get(f"/api/projects?workspace_id={workspace_id}", headers=editor_headers)
        assert project_list.status_code == 200
        assert project_list.json()[0]["id"] == project["id"]
        updated = client.put(f"/api/projects/{project['id']}", headers=editor_headers, json={
            "expected_version": 1, "title": "新型储能产业链", "state": {**project["state"], "projectTitle": "新型储能产业链"},
        })
        assert updated.status_code == 200
        assert updated.json()["version"] == 2
        conflict = client.put(f"/api/projects/{project['id']}", headers=owner_headers, json={"expected_version": 1, "title": "过期修改"})
        assert conflict.status_code == 409
        assert conflict.json()["detail"]["code"] == "project_conflict"
        assert conflict.json()["detail"]["current_version"] == 2
        assert client.get(f"/api/projects/{project['id']}", headers=owner_headers).json()["title"] == "新型储能产业链"

        changed = client.patch(f"/api/workspaces/{workspace_id}/members/{editor_id}", headers=owner_headers, json={"role": "viewer"})
        assert changed.status_code == 200
        assert changed.json()["role"] == "viewer"
        assert client.put(f"/api/projects/{project['id']}", headers=editor_headers, json={"expected_version": 2, "title": "无权修改"}).status_code == 403
        assert client.delete(f"/api/workspaces/{workspace_id}/members/{owner_session['user']['id']}", headers=owner_headers).status_code == 409
        assert client.delete(f"/api/workspaces/{workspace_id}/members/{editor_id}", headers=owner_headers).status_code == 204
        assert client.get(f"/api/workspaces/{workspace_id}/members", headers=editor_headers).status_code == 403
        assert client.delete(f"/api/projects/{project['id']}", headers=owner_headers).status_code == 204
        assert client.get(f"/api/projects/{project['id']}", headers=owner_headers).status_code == 404


def test_wechat_login_creates_and_reuses_account(monkeypatch) -> None:
    class FakeResponse:
        def raise_for_status(self) -> None: pass
        def json(self) -> dict[str, str]: return {"openid": "test-wechat-openid", "unionid": "test-unionid"}

    class FakeClient:
        def __init__(self, **_: object) -> None: pass
        async def __aenter__(self): return self
        async def __aexit__(self, *_: object) -> None: pass
        async def get(self, url: str, params: dict[str, str]):
            assert url.endswith("/sns/jscode2session")
            assert params["js_code"] == "valid-test-code"
            return FakeResponse()

    monkeypatch.setattr(main_module, "settings", SimpleNamespace(wechat_app_id="test-app-id", wechat_app_secret="test-app-secret"))
    monkeypatch.setattr(main_module.httpx, "AsyncClient", FakeClient)
    with TestClient(app) as client:
        registered = client.post("/api/auth/register", json={"email": "wechat-bind@example.com", "password": "safe-password-123", "display_name": "邮箱研究员", "workspace_name": "绑定测试空间"})
        assert registered.status_code == 201
        bound = client.post("/api/auth/bind-wechat", headers={"Authorization": f"Bearer {registered.json()['token']}"}, json={"code": "valid-test-code"})
        assert bound.status_code == 200
        assert bound.json()["bound"] is True
        first = client.post("/api/auth/wechat", json={"code": "valid-test-code", "display_name": "微信研究员"})
        second = client.post("/api/auth/wechat", json={"code": "valid-test-code", "display_name": "不会覆盖名称"})
        assert first.status_code == 200
        assert second.status_code == 200
        assert first.json()["user"]["id"] == registered.json()["user"]["id"]
        assert first.json()["user"]["id"] == second.json()["user"]["id"]
        assert first.json()["workspace"]["id"] == second.json()["workspace"]["id"]


def test_finding_citations_are_checked_against_exact_pages() -> None:
    result = validate_findings(
        {
            "company": "示例公司",
            "findings": [
                {"category": "capacity", "title": "产能", "value": "新增产线", "page_numbers": [2], "quote": "计划新增两条产线", "confidence": 0.9},
                {"category": "financial", "title": "收入", "value": "100亿元", "page_numbers": [3], "quote": "并不存在的原文", "confidence": "unknown"},
            ],
        },
        {2: "公司计划新增两条产线，并于下半年投产。", 3: "营业收入保持增长。"},
    )
    assert result["validation"] == {"total": 2, "matched": 1, "unmatched": 1}
    assert result["findings"][0]["citation_status"] == "matched"
    assert result["findings"][1]["citation_status"] == "unmatched"


def test_desktop_ai_settings_are_saved_locally(monkeypatch) -> None:
    from app.config import settings

    settings_path = TEST_DATA_DIR / "desktop-settings.json"
    monkeypatch.setattr(settings, "app_env", "desktop")
    monkeypatch.setattr(settings, "desktop_settings_path", settings_path)
    monkeypatch.setattr(settings, "deepseek_api_key", "")
    monkeypatch.setattr(settings, "deepseek_base_url", "https://api.deepseek.com")
    monkeypatch.setattr(settings, "deepseek_model", "deepseek-v4-flash")
    with TestClient(app) as client:
        before = client.get("/api/local-settings")
        assert before.status_code == 200
        assert before.json()["configured"] is False
        saved = client.put(
            "/api/local-settings",
            json={"api_key": "unit-test-key", "base_url": "https://api.deepseek.com", "model": "deepseek-chat"},
        )
        assert saved.status_code == 200
        assert saved.json() == {"configured": True, "base_url": "https://api.deepseek.com", "model": "deepseek-chat"}
    assert '"api_key": "unit-test-key"' in settings_path.read_text(encoding="utf-8")
